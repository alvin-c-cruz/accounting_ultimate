from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, current_app, send_file
from io import BytesIO
import json
from sqlalchemy.exc import IntegrityError
from .models import Account as Obj
from .models import ObjUser as Preparer
from .models import ObjAdmin as Approver
from .forms import Form
from application.extensions import db
from application.blueprints.user import login_required, roles_accepted
from flask_login import current_user
import openpyxl
from werkzeug.utils import secure_filename
import os

from . import app_name, app_label


bp = Blueprint(app_name, __name__, template_folder="pages", url_prefix=f"/{app_name}")
ROLES_ACCEPTED = app_label


@bp.route("/")
@login_required
@roles_accepted([ROLES_ACCEPTED])
def home():
    rows = Obj.query.order_by(getattr(Obj, f"account_number")).all()

    context = {
        "rows": rows
    }

    return render_template(f"{app_name}/home.html", **context)


@bp.route("/add", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def add():
    if request.method == "POST":
        form = Form()
        form._post(request.form, current_user.id)

        if form._validate_on_submit():
            form._save()
            flash(f"{form.account_title} has been saved.")
            return redirect(url_for(f'{app_name}.add'))
    else:
        form = Form()

    context = {
        "form": form,
    }

    return render_template(f"{app_name}/form.html", **context)


@bp.route(f"/edit/<int:record_id>", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def edit(record_id):   
    if request.method == "POST":
        form = Form()
        form._post(request.form, current_user.id)

        if form._validate_on_submit():
            form._save()
            return redirect(url_for(f'{app_name}.home'))

    else:
        obj = Obj.query.get(record_id)
        form = Form()
        form._populate(obj)

    context = {
        "form": form,
    }

    return render_template(f"{app_name}/form.html", **context)


@bp.route("/delete/<int:record_id>", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def delete(record_id):   
    obj = Obj.query.get_or_404(record_id)
    preparer = obj.preparer
    try:
        db.session.delete(preparer)
        db.session.delete(obj)
        db.session.commit()
        flash(f"{obj} has been deleted.", category="success")
    except IntegrityError:
        db.session.rollback()
        flash(f"Cannot delete {obj} because it has related records.", category="error")

    return redirect(url_for(f'{app_name}.home'))


@bp.route("/approve/<int:record_id>", methods=['GET'])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def approve(record_id):
    if not current_user.admin:
        flash("Administrator rights required.", category="error")
        return redirect(url_for(f"{app_name}.home"))
    
    obj = Obj.query.get_or_404(record_id)

    data = {
        f"{app_name}_id": record_id,
        "user_id": current_user.id
    }

    approve = Approver(**data)

    db.session.add(approve)
    db.session.commit()

    flash(f"Approved: {getattr(obj, f"{app_name}_name")}", category="success")
    return redirect(url_for(f"{app_name}.home"))   
    

@bp.route("/unlock/<int:record_id>", methods=['GET'])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def unlock(record_id):
    if not current_user.admin:
        flash("Administrator rights required.", category="error")
        return redirect(url_for(f"{app_name}.home"))
    
    obj = Obj.query.get_or_404(record_id)

    data = {
        f"{app_name}_id": record_id,
    }

    approve = Approver.query.filter_by(**data).first()
    
    db.session.delete(approve)
    db.session.commit()

    flash(f"Unlocked: {getattr(obj, f"{app_name}_name")}", category="error")
    return redirect(url_for(f"{app_name}.home"))   
    

@bp.route("/autocomplete", methods=['GET'])
@login_required
def _autocomplete():
    options = [f"{i.account_number}: {i.account_title}" for i in Obj.query.order_by("account_number").all()]
    return Response(json.dumps(options), mimetype='application/json')


@bp.route("/upload", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def upload():
    file = request.files.get("xlsx_file")

    temp_dir = os.path.join(current_app.instance_path, "temp")
    os.makedirs(temp_dir, exist_ok=True)

    # Clean old files
    for f in os.listdir(temp_dir):
        os.remove(os.path.join(temp_dir, f))

    if not file or not file.filename.endswith(".xlsx"):
        flash("Please upload a valid .xlsx file.", "danger")
        return redirect(url_for(f"{app_name}.home"))

    filepath = os.path.join(temp_dir, secure_filename(file.filename))
    file.save(filepath)

    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active

        imported = 0
        skipped = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            account_number, account_title, account_description = row[:3]

            if not account_number or not account_title:
                continue

            existing = Obj.query.filter(
                (Obj.account_number == str(account_number)) |
                (Obj.account_title == str(account_title))
            ).first()

            if existing:
                skipped += 1
                continue

            account = Obj(
                account_number=str(account_number).upper(),
                account_title=str(account_title).upper(),
                account_description=str(account_description or "").upper()
            )

            db.session.add(account)
            db.session.commit()

            preparer_data = {
                f"{app_name}_id": account.id,
                "user_id": current_user.id
            }
            preparer = Preparer(**preparer_data)
            db.session.add(preparer)
            db.session.commit()

            imported += 1

        flash(f"{imported} record(s) imported successfully. {skipped} skipped due to duplicates.", "success")

    except Exception as e:
        flash(f"Error processing file: {str(e)}", "danger")

    return redirect(url_for(f"{app_name}.home"))


@bp.route("/download-template", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chart of Accounts"

    # Header row
    ws.append(["Account Number", "Account Title", "Description"])

    # Optionally, add sample data
    # ws.append(["101", "Cash", "Cash on hand"])

    # Save workbook to memory
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="account.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
