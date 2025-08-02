import os
from flask import g
from pandas import DataFrame, concat
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side

from .models import Disbursement, DisbursementDetail
from .. vendor import Vendor  # adjust path as needed
from collections import defaultdict
from decimal import Decimal
from openpyxl.utils import get_column_letter

# from .. account_type import AccountType
from .. account import Account

from . import app_name, app_label

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

double_rule_border = Border(bottom=Side(style='double'))

ALIGNMENT = {
                "Date": Alignment(horizontal="center", vertical="top"),
                "No.": Alignment(horizontal="center", vertical="top"),
                "Invoice Number": Alignment(horizontal="center", vertical="top"),
                "Vendor": Alignment(horizontal="left", vertical="top", wrap_text=True),
                "Particulars": Alignment(horizontal="left", vertical="top", wrap_text=True),
                "Posted": Alignment(horizontal="center", vertical="center", wrap_text=True)
            }

NUMBER_FORMAT = {
                "Date": "yyyy-mmm-dd",
                "No.": "General",
                "Invoice Number": "General",
                "Vendor": "General",
                "Particulars": "General",
                "Posted": "General",
            }

COLUMN_WIDTH = {
                "Date": 12,
                "No.": 10,
                "Invoice Number": 12,
                "Vendor": 20,
                "Particulars": 25,
                "Posted": 7,
            }


def create_journal(data, app, date_from, date_to):
    list_files = os.listdir(os.path.join(app.instance_path, "temp"))
    for file in list_files:
        os.remove(os.path.join(app.instance_path, "temp", file))

    filename = os.path.join(app.instance_path, "temp", f"{app_name}.xlsx")

    wb = Workbook()

    WriteData(wb, data, date_from, date_to)

    wb.save(filename)
    wb.close()

    return filename


def WriteData(wb, data, date_from, date_to):
    ws = wb.active
    ws.title = "Disbursement Journal"

    # Title
    ws.append([f"{app_label.upper()} DISBURSEMENT JOURNAL"])
    ws.append([f"PERIOD COVERED: {date_from} TO {date_to}"])
    ws.append([])

    # --- Determine all account titles to create dynamic columns ---
    account_names = set()
    for disbursement in data:
        for detail in disbursement.disbursement_details:
            if detail.account:
                account_names.add(detail.account.account_title)

    account_names = sorted(account_names)  # keep columns in order
    header = ["Date", "No.", "Invoice Number", "Vendor", "Particulars"] + account_names
    ws.append(header)

    # Style header row
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    totals = defaultdict(Decimal)

    # --- Write data rows ---
    for disbursement in data:
        if not disbursement.cancelled:
            base_info = {
                "date": disbursement.record_date,
                "no": disbursement.record_number,
                "ap": disbursement.ap_number,
                "vendor": disbursement.vendor.vendor_name if disbursement.vendor else "",
                "desc": disbursement.description
            }
        else:
            base_info = {
                "date": disbursement.record_date,
                "no": disbursement.record_number,
                "ap": "",
                "vendor": "CANCELLED",
                "desc": ""
            }

        # Aggregate account values for this disbursement
        row_data = defaultdict(Decimal)
        for detail in disbursement.disbursement_details:
            name = detail.account.account_title if detail.account else ""
            if not disbursement.cancelled:
                row_data[name] += Decimal(detail.debit or 0) - Decimal(detail.credit or 0)
                totals[name] += Decimal(detail.debit or 0) - Decimal(detail.credit or 0)
            else:
                row_data[name] += 0
                totals[name] += 0                

        # Prepare the row
        row = [
            base_info["date"],
            base_info["no"],
            base_info["ap"],
            base_info["vendor"],
            base_info["desc"],
        ]

        # Add amounts in account columns
        for acct in account_names:
            amt = row_data.get(acct, Decimal(0))
            row.append(amt if amt != 0 else "")

        ws.append(row)

        # Style the row
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right", vertical="top")
            if isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = '#,##0.00_ ;(#,##0.00)'

    # --- Append Totals ---
    total_row = [""] * 5  # Skip fixed columns
    for acct in account_names:
        amt = totals[acct]
        total_row.append(amt if amt != 0 else "")

    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.border = double_rule_border
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right", vertical="top")
        if isinstance(cell.value, (int, float, Decimal)):
            cell.number_format = '#,##0.00_ ;(#,##0.00)'

    # --- Set column widths ---
    col_widths = {
        "Date": 12,
        "No.": 10,
        "Invoice Number": 15,
        "Vendor": 30,
        "Particulars": 25
    }

    for i, col_name in enumerate(header, start=1):
        width = col_widths.get(col_name, 15)
        ws.column_dimensions[get_column_letter(i)].width = width