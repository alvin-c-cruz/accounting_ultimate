import os
from flask import g
from pandas import DataFrame, concat
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from datetime import datetime

from .models import DisbursementExtra, DisbursementExtraDetail
from ... register.vendor import Vendor  # adjust path as needed
from collections import defaultdict
from decimal import Decimal
from openpyxl.utils import get_column_letter

# from .. account_type import AccountType
from ... account import Account

from . import app_name, app_label

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

double_rule_border = Border(bottom=Side(style='double'))

ALIGNMENT = {
                "Date": Alignment(horizontal="center", vertical="top"),
                "CD Extra No.": Alignment(horizontal="center", vertical="top"),
                "AP Extra No.": Alignment(horizontal="center", vertical="top"),
                "Vendor": Alignment(horizontal="left", vertical="top", wrap_text=True),
                "Particulars": Alignment(horizontal="left", vertical="top", wrap_text=True),
            }

NUMBER_FORMAT = {
                "Date": "yyyy-mmm-dd",
                "CD Extra No.": "General",
                "AP Extra No.": "General",
                "Vendor": "General",
                "Particulars": "General",
            }

COLUMN_WIDTH = {
                "Date": 12,
                "CD Extra No.": 10,
                "AP Extra No.": 12,
                "Vendor": 20,
                "Particulars": 25,
            }


def create_journal(data, app, date_from, date_to):
    list_files = os.listdir(os.path.join(app.instance_path, "temp"))
    for file in list_files:
        os.remove(os.path.join(app.instance_path, "temp", file))

    filename = os.path.join(app.instance_path, "temp", f"{app_name}.xlsx")

    wb = Workbook()

    WriteData(wb, data, date_from, date_to)

    wb.save(filename)
    wb.close()

    return filename


def WriteData(wb, data, date_from, date_to):
    ws = wb.active
    ws.title = app_name
    date_from = datetime.strptime(date_from, "%Y-%m-%d").strftime("%B %d, %Y")
    date_to = datetime.strptime(date_to, "%Y-%m-%d").strftime("%B %d, %Y")

    # Title
    ws.append([f"{app_label.upper()} JOURNAL"])
    ws.append([f"From {date_from} to {date_to}"])
    ws.append([])
    
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"].font = Font(bold=True)

    # --- Determine all account titles to create dynamic columns ---
    account_names = set()
    for disbursement in data:
        for detail in disbursement.disbursement_extra_details:
            if detail.account:
                account_names.add(detail.account.account_title)

    account_names = sorted(account_names)  # keep columns in order
    header = ["Date", "CD Extra No.", "AP Extra No.", "Vendor", "Particulars"] + account_names
    ws.append(header)

    # Style header row
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    totals = defaultdict(Decimal)

    # --- Write data rows ---
    for disbursement in data:
        if not disbursement.cancelled:
            base_info = {
                "date": disbursement.record_date,
                "no": disbursement.record_number,
                "ap": disbursement.ap_number,
                "vendor": disbursement.vendor.vendor_name if disbursement.vendor else "",
                "desc": disbursement.description,
            }
        else:
            base_info = {
                "date": disbursement.record_date,
                "no": disbursement.record_number,
                "ap": "",
                "vendor": "CANCELLED",
                "desc": "",
            }

        # Aggregate account values for this disbursement
        row_data = defaultdict(Decimal)
        for detail in disbursement.disbursement_extra_details:
            name = detail.account.account_title if detail.account else ""
            if not disbursement.cancelled:
                row_data[name] += Decimal(detail.debit or 0) - Decimal(detail.credit or 0)
                totals[name] += Decimal(detail.debit or 0) - Decimal(detail.credit or 0)
            else:
                row_data[name] += 0
                totals[name] += 0                

        # Prepare the row
        row = [
            base_info["date"],
            base_info["no"],
            base_info["ap"],
            base_info["vendor"],
            base_info["desc"],
        ]

        # Add amounts in account columns
        for acct in account_names:
            amt = row_data.get(acct, Decimal(0))
            row.append(amt if amt != 0 else "")

        ws.append(row)

        # Style the row
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            col_letter = get_column_letter(cell.column)
            if col_letter in ("A", "B", "C", "I"):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_letter in ("D", "E"):
                cell.alignment = Alignment(horizontal="left", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="top")
            if isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = '#,##0.00_ ;(#,##0.00)'

    # --- Append Totals ---
    # Determine row bounds for data
    data_start_row = 5  # Header is on row 4; data starts at 5
    data_end_row = ws.max_row

    # Start building the total row
    total_row = [""] * 5  # Leave first 5 columns blank
    for col_idx in range(6, 6 + len(account_names)):  # starting from column F (index 6)
        col_letter = get_column_letter(col_idx)
        formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
        total_row.append(formula)

    ws.append(total_row)

    # Style the total row
    for cell in ws[ws.max_row]:
        cell.border = double_rule_border
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right", vertical="top")
        cell.number_format = '#,##0.00_ ;(#,##0.00)'

    for cell in ws[ws.max_row]:
        cell.border = double_rule_border
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right", vertical="top")
        if isinstance(cell.value, (int, float, Decimal)):
            cell.number_format = '#,##0.00_ ;(#,##0.00)'

    # --- Set column widths ---
    col_widths = {
        "Date": 12,
        "CD Extra No.": 10,
        "AP Extra No.": 10,
        "Vendor": 30,
        "Particulars": 25,
    }

    for i, col_name in enumerate(header, start=1):
        width = col_widths.get(col_name, 15)
        ws.column_dimensions[get_column_letter(i)].width = width