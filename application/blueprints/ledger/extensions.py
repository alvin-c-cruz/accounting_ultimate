from openpyxl import Workbook
from pathlib import Path
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from flask import url_for

from .. account import Account
from .. books_of_accounts.sales import Sales
from .. books_of_accounts_extra.sales_extra import SalesExtra
from .. books_of_accounts.receipt import Receipt
from .. books_of_accounts_extra.receipt_extra import ReceiptExtra
from .. books_of_accounts.accounts_payable import AccountsPayable
from .. books_of_accounts_extra.accounts_payable_extra import AccountsPayableExtra
from .. books_of_accounts.disbursement import Disbursement
from .. books_of_accounts_extra.disbursement_extra import DisbursementExtra
from .. books_of_accounts.general import General
from .. books_of_accounts_extra.general_extra import GeneralExtra


font_14_bold = Font(size=14, bold=True)
font_10_bold = Font(size=10, bold=True)
font_10 = Font(size=10)


thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

align_center_top = Alignment(horizontal="center", vertical="top")
align_center_center = Alignment(horizontal="center", vertical="center")
align_center_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left_top = Alignment(horizontal="left", vertical="top")
align_right_top = Alignment(horizontal="right", vertical="top")
align_left_top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)


module_name = {
    Sales: "sales", 
    SalesExtra: "sales_extra", 
    Receipt: "receipt", 
    ReceiptExtra: "receipts_extra",
    AccountsPayable: "accounts_payable",
    AccountsPayableExtra: "accounts_payable_extra",
    Disbursement: "disbursement",
    DisbursementExtra: "disbursements_extra",
    General: "general",
    GeneralExtra: "general_extra",
}

book = {
    Sales: "SJ-Corp", 
    SalesExtra: "SJ-Extra", 
    Receipt: "CR-Corp", 
    ReceiptExtra: "CR-Extra",
    AccountsPayable: "AP-Corp",
    AccountsPayableExtra: "APJ-Extra",
    Disbursement: "CD-Corp",
    DisbursementExtra: "CD-Extra",
    General: "GJ",
    GeneralExtra: "GJ-Extra"
}

ref_num = {
    Sales: "record_number", 
    SalesExtra: "record_number", 
    Receipt: "record_number", 
    ReceiptExtra: "record_number",
    AccountsPayable: "record_number",
    AccountsPayableExtra: "record_number",
    Disbursement: "record_number",
    DisbursementExtra: "record_number",
    General: "record_number",
    GeneralExtra: "record_number",
}

def beginning_balance(account, date_from):
    _beginning = 0
    # Beginning
    for obj in (
                Sales, 
                SalesExtra, 
                Receipt, 
                ReceiptExtra,
                AccountsPayable,
                AccountsPayableExtra,
                Disbursement,
                DisbursementExtra,
                General,
                GeneralExtra
                ):
        vouchers = obj.query.filter(obj.record_date<date_from).all()
        for voucher in vouchers:
            for entry in voucher.entries:
                if entry.account == account:
                    _beginning += entry.debit - entry.credit
    
    return _beginning


def account_transactions(account, date_to, date_from):
    # Transactions
    _transactions = []
    
    for obj in (
                Sales, 
                SalesExtra, 
                Receipt, 
                ReceiptExtra,
                AccountsPayable,
                AccountsPayableExtra,
                Disbursement,
                DisbursementExtra,
                General,
                GeneralExtra
                ):
        vouchers = obj.query.filter(obj.record_date>=date_from, obj.record_date<=date_to).all()
        for voucher in vouchers:
            for entry in voucher.entries:
                if entry.account == account:
                    _transaction = {
                            "record_id": voucher.id,
                            "date": voucher.record_date,
                            "book": book[obj],
                            "reference": getattr(voucher, ref_num[obj]),
                            "particulars": voucher.notes,
                            "debit": entry.debit,
                            "credit": entry.credit,
                            "formatted_debit": '{:,.2f}'.format(entry.debit),
                            "formatted_credit": '{:,.2f}'.format(entry.credit),
                            "date_posted": voucher.date_posted if hasattr(voucher, "date_posted") else "",
                            "edit_url": url_for(module_name[obj] + ".edit", id=voucher.id),
                            "view_url": url_for(module_name[obj] + ".view", id=voucher.id)  if hasattr(voucher, "date_posted") else "",
                        }
                        
                    if obj in (Sales,SalesExtra, Receipt, ReceiptExtra,):
                        description = voucher.customer.customer_name
                    elif obj in (AccountsPayable, AccountsPayableExtra, Disbursement, DisbursementExtra, General, GeneralExtra):
                        description = voucher.vendor.vendor_name
                    else:
                        description = ""

                    _transaction["description"] = description
                        
                    if obj in (Disbursement, DisbursementExtra):
                        check_number = voucher.check_number
                    else:
                        check_number = ""
                            
                    _transaction["check_number"] = check_number
                            
                    _transactions.append(_transaction)

    return sorted(_transactions, key=lambda x: (x['date'], x['book']))


def write_trial_balance_header(ws):
    row_num = 1
    
    columns = {
        "A": "Account Number",
        "B": "Account Title",
        "C": "Beginning",
        "D": "Debit",
        "E": "Credit",
        "F": "Balance"
    }
    
    for column, label in columns.items():
        cell = ws[f"{column}{row_num}"]
        cell.value = label
    
    row_num += 1
    
    return row_num

def create_ledger_all(accounts, date_from, date_to, instance_path):
    root = Path(instance_path)
    filename = root / "temp" / "general_ledger.xlsx"

    wb = Workbook()    
    ws_trial_balance = wb["Sheet"]
    ws_trial_balance.title = "Trial Balance"
    wtb_row_num = write_trial_balance_header(ws_trial_balance)
    
    for account in accounts:
        account_number = account.account_number
        wb.create_sheet(account_number)
        ws = wb[account_number]
        
        beginning = beginning_balance(account, date_from)
        transactions = account_transactions(account, date_to, date_from)

        row_num = 1
        row_num = write_header(ws, row_num, account, date_from, date_to)
        row_num = write_columns(ws, row_num)
        row_num = write_beginning_balance(ws, beginning, row_num)
        row_num = write_transactions(ws, transactions, row_num)
        
        ws.auto_filter.ref = f"A7:L{row_num}"
        ws.freeze_panes = 'E8'
        
        wtb_row_num = write_trial_balance(ws_trial_balance, wtb_row_num, ws, row_num, account)
        
    
    wb.save(filename)
    wb.close()

    return filename


def create_ledger(account, beginning, transactions, date_from, date_to, instance_path):
    root = Path(instance_path)
    filename = root / "temp" / "general_ledger.xlsx"

    wb = Workbook()
    ws = wb["Sheet"]
    ws.title = account.account_number

    row_num = 1
    row_num = write_header(ws, row_num, account, date_from, date_to)
    row_num = write_columns(ws, row_num)
    row_num = write_beginning_balance(ws, beginning, row_num)
    row_num = write_transactions(ws, transactions, row_num)

    ws.auto_filter.ref = f"A7:L{row_num}"
    ws.freeze_panes = 'E8'

    wb.save(filename)
    wb.close()

    return filename


def write_header(ws, row_num, account, date_from, date_to):
    cell = ws[f"A{row_num}"]
    cell.value = "Rowell Industrial Corporation"
    cell.font = font_14_bold
    row_num += 1

    cell = ws[f"A{row_num}"]
    cell.value = f"{account.account_number}: {account.account_title}"
    cell.font = font_14_bold
    row_num += 1

    cell = ws[f"A{row_num}"]
    cell.value = "General Ledger"
    cell.font = font_14_bold
    row_num += 1

    cell = ws[f"A{row_num}"]
    cell.value = f"From {date_from.strftime('%B %d, %Y')} to {date_to.strftime('%B %d, %Y')}"
    cell.font = font_10
    row_num += 3

    return row_num


def write_trial_balance(ws_trial_balance, wtb_row_num, ws, account_row_num, account):
    ws = ws_trial_balance
    
    columns = {
        "A": account.account_number,
        "B": account.account_title,
        "C": f"='{account.account_number}'!I8",
        "D": f"=SUM('{account.account_number}'!G:G)",
        "E": f"=SUM('{account.account_number}'!H:H)",
        "F": f"=C{wtb_row_num}+D{wtb_row_num}-E{wtb_row_num}",
        "G": f"='{account.account_number}'!I{account_row_num-1}",
        "H": f"=F{wtb_row_num}-G{wtb_row_num}"
    }
    
    for column, label in columns.items():
        cell = ws_trial_balance[f"{column}{wtb_row_num}"]
        cell.value = label
    
    wtb_row_num += 1
    
    print(f"Written {account.account_title}")
    
    return wtb_row_num
    
    

def write_columns(ws, row_num):
    columns = {
        "A": "Voucher Date",
        "B": "Bank Date",
        "C": "Book",
        "D": "Reference",
        "E": "Check No.",
        "F": "Name",
        "G": "Particulars",
        "H": "Debit",
        "I": "Credit",
        "J": "Running Balance",
        "K": "POSTED",
        "L": "Go to voucher",
    }

    column_width = {
        "A": 12,
        "B": 9,
        "C": 9,
        "D": 9,
        "E": 9,
        "F": 35,
        "G": 40,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 9,
        "L": 9,
    }

    for column_letter, value in columns.items():
        cell = ws[f"{column_letter}{row_num}"]
        cell.value = value
        cell.font = font_10_bold
        cell.border = thin_border
        cell.alignment = align_center_center_wrap
        
        ws.column_dimensions[column_letter].width = column_width[column_letter]

    row_num += 1

    return row_num


def write_beginning_balance(ws, beginning_balance, row_num):
    columns = {
        "A": "",
        "B": "",
        "C": "",
        "D": "",
        "E": "",
        "F": "",
        "G": "Beginning Balance",
        "H": "",
        "I": "",
        "J": beginning_balance,
        "K": "",
        "L": "",
    }
    for column_letter, value in columns.items():
        cell = ws[f"{column_letter}{row_num}"]
        cell.value = value
        cell.font = font_10
        cell.border = thin_border
        
        if column_letter == "J": 
            cell.number_format = "#,##0.00"
            cell.alignment = align_right_top

    row_num += 1

    return row_num


def write_transactions(ws, transactions, row_num):
    for transaction in transactions:
        columns = {
            "A": transaction["date"],
            "B": transaction["bank_date"],
            "C": transaction["book"],
            "D": transaction["reference"],
            "E": transaction["check_number"],
            "F": transaction["description"],
            "G": transaction["particulars"],
            "H": transaction["debit"],
            "I": transaction["credit"],
            "J": '=INDIRECT("J"&ROW()-1)+INDIRECT("H"&ROW())-INDIRECT("I"&ROW())',
            "K": "Yes" if transaction["date_posted"] else "No",
            "L": "click here" 
        }
        for column_letter, value in columns.items():
            cell = ws[f"{column_letter}{row_num}"]
            cell.value = value
            cell.font = font_10
            cell.border = thin_border

            if column_letter in ("A","B"):
                cell.alignment = align_right_top
                cell.number_format = "dd-mmm-yyyy"
            elif column_letter in ("C", "D", "E"):
                cell.alignment = align_center_top
            elif column_letter in ("F", "G"):
                cell.alignment = align_left_top
            elif column_letter in ("H", "I", "J"):
                cell.alignment = align_right_top
                cell.number_format = "#,##0.00"

            if column_letter in ("D", "L"):
                if transaction["date_posted"]:
                    cell.hyperlink = "https://jasmincy11.pythonanywhere.com" + transaction["view_url"]
                else:
                    cell.hyperlink = "https://jasmincy11.pythonanywhere.com" + transaction["edit_url"]

        row_num += 1
    
    return row_num
